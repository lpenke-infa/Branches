from openpyxl.styles import numbers
from create_pivot_table import *
from datetime import datetime
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from data import *
import openpyxl
import time
import os

# get coulmn name by column number
def getExcelColumnLabel(column_number):
    result = ''
    while column_number > 0:
        remainder = (column_number - 1) % 26
        result = chr(ord('A') + remainder) + result
        column_number = (column_number - 1) // 26
    return result

# get month, year - 'Dec 2024'
def getMonthYear(date):
    month_dict = {1: 'Jan',2: 'Feb',3: 'Mar',4: 'Apr',5: 'May',6: 'Jun',7: 'Jul',8: 'Aug',9: 'Sep',10: 'Oct',11: 'Nov',12: 'Dec'}
    month = date.month
    if(date.day >= 30):
        month += 1
    if(month==13):
        return month_dict[1]+ " " + str(date.year+1)
    else:
        return month_dict[month]+ " " + str(date.year)

def getDataFromCSV(filename='', data=[], isMasterOrg=False):
    # print(data)
    # return 0
    file = open(filename, 'r')
    lines = file.readlines()
    # lines.extend(data)
    final_data = []
    if(isMasterOrg):
        for line in data:
            date = datetime.strptime(line[0], '%Y-%m-%d %H:%M:%S.0')
            final_data.append(([line[0], getMonthYear(date), line[1], line[2], float(line[3])]))
    for line in lines[1:]:
        line = line.split(',')
        date = datetime.strptime(line[0], '%Y-%m-%d %H:%M:%S.0')
        if(isMasterOrg):
            final_data.append([line[0], getMonthYear(date), line[1], line[2], float(line[3])])
        else:
            final_data.append([line[0], getMonthYear(date), line[1], float(line[4]), float(line[5])])
    final_data = sorted(final_data)
    final_data.insert(0, header)
    return final_data

def subOrgConsumption(filename):
    workbook = openpyxl.load_workbook(filename)
    subOrgConsumption = dict()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        subOrgConsumption[sheet_name] = []
        temp = []
        for row in sheet.iter_rows(values_only=True):
            temp.append(list(row))
        subOrgConsumption[sheet_name] = temp
    workbook.close()
    
    return subOrgConsumption

date_style = NamedStyle(name='custom_date', number_format='mmm-yy')
def writeDataToExcel(sheet='', data=[], isMasterOrg=False, tgtWorkbook=''):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])
    workbook.create_sheet(sheet)
    worksheet = workbook[sheet]
    tgtWorkbook.create_sheet(orgList[sheet])
    tgtWorksheet = tgtWorkbook[orgList[sheet]]
    for row in data:
        worksheet.append(row)
        tgtWorksheet.append(row)

    if(isMasterOrg):
        workbook.save(targetPath+'IPU_Report_'+ orgList[sheet].split('_')[-1] + '_All_Org_Level_'+ date +'.xlsx')
    else:
        fname = filenamesToSave[sheet].split('-')
        workbook.save(targetPath+'IPU_Report_'+ '_'.join(fname) + '_' + date +'.xlsx')
    workbook.close()

def index():
    filenames = os.listdir(sourcePath)
    tgtWorkbook = openpyxl.Workbook()
    tgtWorkbook.remove(tgtWorkbook['Sheet'])

    for specialSheet, masterOrg, masterOrgs_, subOrgs in specialOrgs:
        tgtWorkbook.create_sheet(specialSheet)
    
    for masterOrg in masterOrgs:
        fname = orgList[masterOrg].split('_')[1:]
        tgtWorkbook.create_sheet('_'.join(fname)+' - Summary Reports')
        
    subOrgConsumptionData = subOrgConsumption(subOrgConsumptionFileName)
    
    # print(subOrgConsumptionData)
    
    for filename in filenames:
        sheet = filename[:31]
        if(sheet in masterOrgs):
            data = getDataFromCSV(filename=sourcePath+filename, data=subOrgConsumptionData[sheet], isMasterOrg=True)
            writeDataToExcel(sheet=sheet, data=data, isMasterOrg=True,tgtWorkbook=tgtWorkbook)
        else:
            data = getDataFromCSV(filename=sourcePath+filename, isMasterOrg=False)
            writeDataToExcel(sheet=sheet, data=data, isMasterOrg=False,tgtWorkbook=tgtWorkbook)
    
    for specialSheet, masterOrg, masterOrgs_, subOrgs in specialOrgs:
        data = []
        tgtWorkbook.create_sheet(masterOrg)
        for org in masterOrgs_:
            sheet = tgtWorkbook[orgList[org]]
            for row in sheet.iter_rows(values_only=True):
                # print(row)
                if(row[2]=='Sub-Organization'):
                    continue
                data.append(row)
        if(masterOrg in subOrgConsumptionData):
            for line in subOrgConsumptionData[masterOrg]:
                date = datetime.strptime(line[0], '%Y-%m-%dT%H:%M:%SZ')
                data.append(([line[0], getMonthYear(date), line[1], line[2], float(line[3])]))
        tgtWorksheet = tgtWorkbook[masterOrg]
        for row in data:
            tgtWorksheet.append(row)

    # Get Months in Each MasterOrg
    monthsInSheet = {}
    j = 0
    for specialSheet, masterOrg, masterOrgs_, subOrgs in specialOrgs:
        lst = []
        sheet = masterOrg
        sheetMonths = tgtWorkbook[sheet]['B']
        sheetOrgs = tgtWorkbook[sheet]['C']
        l = len(sheetMonths)
        for i in range(l):
            if((sheetMonths[i].value not in lst) and (sheetOrgs[i].value in subOrgs)):
                lst.append(sheetMonths[i].value)
        monthsInSheet[specialSheet] = lst
        j += 1
        
    for masterOrg in masterOrgs:
        lst = []
        sheet = orgList[masterOrg]
        sheet = tgtWorkbook[sheet]['B']
        for cell in sheet[1:]:
            if(cell.value not in lst):
                lst.append(cell.value)
        monthsInSheet[masterOrg] = lst

    tgtWorkbook.save(targetPath+tgtFilename)
    tgtWorkbook.close()

    # Yearly Pivot Table
    print('Creating Yearly Sum Pivot Table')
    i = 0
    for specialSheet, masterOrg, masterOrgs_, subOrgs in specialOrgs:
        print('   - '+specialSheet)
        pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=masterOrg, summarySheet=specialSheet, tableName=specialSheet+' table')
        createPivotTable(pt, subOrgs,monthsInSheet[specialSheet])
        i += 1

    for masterOrg in masterOrgs:
        summarySheet = orgList[masterOrg].split('_')[1:]
        summarySheet = '_'.join(summarySheet)+' - Summary Reports'
        print('   - '+summarySheet)
        pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=orgList[masterOrg], summarySheet=summarySheet, tableName=summarySheet+' table')
        createPivotTable(pt=pt,months=monthsInSheet[masterOrg])

    time.sleep(3)

    # Monthly Pivot Table
    print('Creating Monthly Sum Pivot Table')
    for specialSheet, masterOrg, masterOrgs_, subOrgs in specialOrgs:
        print('   - '+specialSheet)
        
        columnNumber = 5
        for month in monthsInSheet[specialSheet]:
            print(f"     - {month}: {getExcelColumnLabel(columnNumber)}")
            
            sumPivotTableName = "SumOfTable"+str(columnNumber)
            columnLabel = getExcelColumnLabel(columnNumber)
            
            # Table 1 - Sum of IPUs Consumed by All Sub Orgs
            pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=masterOrg, summarySheet = specialSheet, cellNumber=columnLabel+"10", tableName=sumPivotTableName)
            worksheet.Cells(1, 1).Offset(7, columnNumber).Value = month
            worksheet.Cells(1, 1).Offset(7, columnNumber).Font.Bold = True
            worksheet.Cells(1, 1).Offset(7, columnNumber).Font.Size = 14
            createPivotTableSubOrg(pt,-4157, month, subOrgs)
            
            pivot_table_data = getPivotTableData(targetPath+tgtFilename, specialSheet, sumPivotTableName)
            
            rowNumber = len(pivot_table_data) + 4 + 10
            worksheet.Cells(1, 1).Offset(rowNumber-5, columnNumber+1).Interior.ColorIndex = 40
            
            # Table [2:-2] - Sum of IPUs Consumed by Each Sub Org for all Services
            for i in pivot_table_data[1:-1]:     
                print('       - '+i[0][0])
                # print(i[0])
                if(i[0][1] > 0 and i[0][0] not in exceptionOrgs):
                    pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=i[0][0], summarySheet = specialSheet, cellNumber=columnLabel+str(rowNumber), tableName=sumPivotTableName+str(rowNumber))
                    worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Value = i[0][0]
                    # worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Interior.ColorIndex = 37
                    worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Font.Bold = True
                    createPivotTableSubOrg(pt,-4157, month)
                    rowNumber += len(getPivotTableData(targetPath+tgtFilename, specialSheet, sumPivotTableName+str(rowNumber))) + 4
                    worksheet.Cells(1, 1).Offset(rowNumber-5, columnNumber+1).Interior.ColorIndex = 40
                elif(i[0][0]=='Sub-Organization'):
                    mergeRange = worksheet.Range(worksheet.Cells(5, columnNumber), worksheet.Cells(5, columnNumber+1))
                    mergeRange.Merge()
                    # print("Done")
                    time.sleep(1)
                    worksheet.Cells(1, 1).Offset(5, columnNumber).Value = f"{int(i[0][1])//6} SubOrgs * 6 IPUs = {int(i[0][1])} IPUs"
                    worksheet.Cells(1, 1).Offset(5, columnNumber).Interior.ColorIndex = 6
                    # worksheet.Cells(1, 1).Offset(5, columnNumber).Font.Bold = True
                    # worksheet.Cells(1, 1).Offset(5, columnNumber).Font.Size = 11
                    # print(i[0][1], '-------')
            
            # # Table -1 - Avg of IPUs Consumed by All Sub Orgs
            # pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=masterOrg, summarySheet = specialSheet, cellNumber=columnLabel+str(rowNumber), tableName="AvgOfTable"+str(columnNumber))
            # worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Value = "Average of "+month
            # worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Font.Bold = True
            # createPivotTableSubOrg(pt,-4106, month, subOrgs)
            # rowNumber += len(getPivotTableData(targetPath+tgtFilename, specialSheet, "AvgOfTable"+str(columnNumber))) + 4
            # worksheet.Cells(1, 1).Offset(rowNumber-5, columnNumber+1).Interior.ColorIndex = 40
            
            columnNumber += 3
            
    for masterOrg in masterOrgs:
        summarySheet = orgList[masterOrg].split('_')[1:]
        summarySheet = '_'.join(summarySheet)+' - Summary Reports'
        print('   - '+summarySheet)
        
        columnNumber = 5
        for month in monthsInSheet[masterOrg]:
            print(f"     - {month}: {getExcelColumnLabel(columnNumber)}")
            
            sumPivotTableName = "SumOfTable"+str(columnNumber)
            columnLabel = getExcelColumnLabel(columnNumber)
            
            # Table 1 - Sum of IPUs Consumed by All Sub Orgs
            pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=orgList[masterOrg], summarySheet = summarySheet, cellNumber=columnLabel+"10", tableName=sumPivotTableName)
            worksheet.Cells(1, 1).Offset(7, columnNumber).Value = month
            worksheet.Cells(1, 1).Offset(7, columnNumber).Font.Bold = True
            worksheet.Cells(1, 1).Offset(7, columnNumber).Font.Size = 14
            createPivotTableSubOrg(pt=pt,numericalRepresentation=-4157, month=month)
            
            pivot_table_data = getPivotTableData(targetPath+tgtFilename, summarySheet, sumPivotTableName)
            
            rowNumber = len(pivot_table_data) + 4 + 10
            worksheet.Cells(1, 1).Offset(rowNumber-5, columnNumber+1).Interior.ColorIndex = 40
            
            # Table [2:-2] - Sum of IPUs Consumed by Each Sub Org for all Services
            for i in pivot_table_data[1:-1]:     
                print('       - '+i[0][0])
                if(i[0][1] > 0 and i[0][0] not in exceptionOrgs):
                    pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=i[0][0], summarySheet = summarySheet, cellNumber=columnLabel+str(rowNumber), tableName=sumPivotTableName+str(rowNumber))
                    worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Value = i[0][0]
                    # worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Interior.ColorIndex = 37
                    worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Font.Bold = True
                    createPivotTableSubOrg(pt=pt,numericalRepresentation=-4157, month=month)
                    rowNumber += len(getPivotTableData(targetPath+tgtFilename, summarySheet, sumPivotTableName+str(rowNumber))) + 4
                    worksheet.Cells(1, 1).Offset(rowNumber-5, columnNumber+1).Interior.ColorIndex = 40
                elif(i[0][0]=='Sub-Organization'):
                    mergeRange = worksheet.Range(worksheet.Cells(5, columnNumber), worksheet.Cells(5, columnNumber+1))
                    mergeRange.Merge()
                    # print("Done")
                    time.sleep(1)
                    worksheet.Cells(1, 1).Offset(5, columnNumber).Value = f"{int(i[0][1])//6} SubOrgs * 6 IPUs = {int(i[0][1])} IPUs"
                    worksheet.Cells(1, 1).Offset(5, columnNumber).Interior.ColorIndex = 6
                    # worksheet.Cells(1, 1).Offset(5, columnNumber).Font.Bold = True
                    # worksheet.Cells(1, 1).Offset(5, columnNumber).Font.Size = 11
                    # print(i[0][1], '-------')
            
            # # Table -1 - Avg of IPUs Consumed by All Sub Orgs
            # pt, worksheet = createPivotTableStyle(path=targetPath+tgtFilename, dataSheet=orgList[masterOrg], summarySheet = summarySheet, cellNumber=columnLabel+str(rowNumber), tableName="AvgOfTable"+str(columnNumber))
            # worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Value = "Average of "+month
            # worksheet.Cells(1, 1).Offset(rowNumber-3, columnNumber).Font.Bold = True
            # createPivotTableSubOrg(pt=pt,numericalRepresentation=-4106, month=month)
            # rowNumber += len(getPivotTableData(targetPath+tgtFilename, summarySheet, "AvgOfTable"+str(columnNumber))) + 4
            # worksheet.Cells(1, 1).Offset(rowNumber-5, columnNumber+1).Interior.ColorIndex = 40
            
            columnNumber += 3

    # xlApp.Quit()

# index()

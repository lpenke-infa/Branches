import openpyxl

def subOrgConsumption(filename):
    workbook = openpyxl.load_workbook(filename)
    subOrgConsumption = dict()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        subOrgConsumption[sheet_name] = []
        temp = []
        for row in sheet.iter_rows(values_only=True):
            temp.append(row)
        subOrgConsumption[sheet_name] = temp
    workbook.close()
    
    return subOrgConsumption

if __name__ == "__main__":
    # Provide the path to your Excel file
    filename = "sub-organization_consumption.xlsx"
    
    # Call the function to read data from each sheet
    subOrgConsumption = subOrgConsumption(filename)
    print(subOrgConsumption)
